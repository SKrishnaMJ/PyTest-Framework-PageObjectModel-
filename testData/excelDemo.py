import openpyxl

book = openpyxl.load_workbook("/Users/saikrishnamj/Desktop/Udemy/pythonDemo.xlsx")
my_dict={}
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Test3":
        for j in range(1,sheet.max_column+1):
            my_dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i, column=j).value
            print(my_dict)

