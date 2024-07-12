import openpyxl

class HomepageData:

    # data = [{"firstname":"Sai Krishna", "email":"saik@gmail.com", "password":"12345", "gender":"Male"},
    #         {"firstname":"Ritika G", "email":"srg001@gmail.com", "password":"87654", "gender":"Female"}]

    @staticmethod
    def getTestData(test_name):
        book = openpyxl.load_workbook("/Users/saikrishnamj/Desktop/Udemy/pythonDemo.xlsx")
        my_dict = {}
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_name:
                for j in range(2, sheet.max_column + 1):
                    my_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [my_dict]
